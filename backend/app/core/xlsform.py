"""XLSForm import — turn a Survey123 / KoBo Collect style .xlsx (a
`survey` sheet, optionally `choices` and `settings`) into GeoCore's own
form shape, so a form built once in the XLSForm workflow doesn't have to
be rebuilt by hand in GeoCore's form builder.

This is a best-effort conversion, **not** a full XLSForm engine — see
docs/CHANGES_XLSFORM_AND_DATA_IMPORT.md for the exact rules. In short:

- `begin group`/`end group` -> a section; `begin repeat`/`end repeat` -> a
  repeatable section. Nested groups (a group inside a group) are
  flattened to one level, since GeoCore's form model is one section deep.
- `geopoint`/`geotrace`/`geoshape` sets the asset type's geometry type
  instead of becoming a form field.
- `select_one`/`select_multiple <list>` pulls options from the `choices`
  sheet.
- `relevant`, `calculation`, and `constraint` are real XLSForm expression
  languages (arbitrary XPath-style expressions) — only common patterns
  are recognized (see `convert_relevant`/`convert_calculation`/
  `convert_constraint` below). Anything else is skipped, not guessed at,
  and reported back in `ParsedForm.warnings` so it can be rebuilt by hand.
- `note`, `start`, `end`, `deviceid`, `audit` questions are dropped
  silently (form implementation detail, not data).
- Unrecognized question types are skipped with a warning.

A field's `field_key` is taken directly from the XLSForm `name` column
(when it's already a valid identifier) rather than being slugified from
the label, because converted `relevant`/`calculation`/`constraint`
expressions reference field keys built from `name` — regenerating keys
from the label would silently break every converted expression.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field as dc_field
from typing import Any

from backend.app.core.slugify import slugify_key


class XLSFormError(ValueError):
    """Raised for anything that stops the import outright — a missing or
    unreadable `survey` sheet, an unparsable file, etc. Individual rows
    that don't convert cleanly don't raise this; they're skipped and
    reported in `ParsedForm.warnings` instead.
    """


# ---------------------------------------------------------------------------
# Result shapes
# ---------------------------------------------------------------------------


@dataclass
class ParsedField:
    label: str
    field_type: str
    field_key: str
    options: list[str] | None = None
    is_required: bool = False
    visibility: dict | None = None
    calculation: str | None = None
    validation: dict | None = None


@dataclass
class ParsedSection:
    title: str
    repeatable: bool = False
    repeat_label: str | None = None
    fields: list[ParsedField] = dc_field(default_factory=list)


@dataclass
class ParsedForm:
    name: str
    geometry_type: str = "point"
    sections: list[ParsedSection] = dc_field(default_factory=list)
    warnings: list[str] = dc_field(default_factory=list)


# ---------------------------------------------------------------------------
# Type mapping
# ---------------------------------------------------------------------------

GEOMETRY_QUESTION_TYPES = {
    "geopoint": "point",
    "geotrace": "line",
    "geoshape": "polygon",
}

# XLSForm `type` -> GeoCore field_type. Keys are matched against the
# lower-cased XLSForm type (or, for select questions, the lower-cased
# first token — "select_one <list>" -> "select_one").
_TYPE_MAP = {
    "text": "text",
    "string": "text",
    "notes": "long_text",
    "integer": "number",
    "decimal": "number",
    "range": "number",
    "date": "date",
    "datetime": "datetime",
    "select_one": "single_select",
    "select_multiple": "multi_select",
    "boolean": "boolean",
    "image": "photo",
    "photo": "photo",
    "video": "video",
    "file": "file",
    "signature": "signature",
}

# Dropped silently — form implementation detail, not data a person would
# expect to see (blueprint / doc: "note, start, end, deviceid, audit").
_DROPPED_TYPES = {"note", "start", "end", "deviceid", "audit"}

_IDENTIFIER_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")


# ---------------------------------------------------------------------------
# relevant / calculation / constraint conversion
# ---------------------------------------------------------------------------

_FIELD_REF_RE = re.compile(r"\$\{([a-zA-Z0-9_]+)\}")
_COND_RE = re.compile(
    r"^\$\{(?P<field>[a-zA-Z0-9_]+)\}\s*(?P<op>!=|>=|<=|=|>|<)\s*(?P<value>.+)$"
)
_SELECTED_RE = re.compile(
    r"^selected\(\s*\$\{(?P<field>[a-zA-Z0-9_]+)\}\s*,\s*(?P<value>.+?)\s*\)$"
)
_CONSTRAINT_RE = re.compile(r"^\.\s*(>=|<=|>|<)\s*(-?\d+(?:\.\d+)?)$")

_COMPARE_OP_MAP = {
    "=": "equals",
    "!=": "not_equals",
    ">": "greater_than",
    "<": "less_than",
    ">=": "greater_or_equal",
    "<=": "less_or_equal",
}

_ALLOWED_CALC_FUNCS = {"round", "abs", "min", "max", "sum"}


def _parse_literal(raw: str) -> Any:
    raw = raw.strip()
    if len(raw) >= 2 and raw[0] == raw[-1] and raw[0] in ("'", '"'):
        return raw[1:-1]
    try:
        if re.fullmatch(r"-?\d+", raw):
            return int(raw)
        return float(raw)
    except ValueError:
        return raw


def _split_combinator(expr: str) -> tuple[list[str] | None, str | None]:
    """Splits on a single combinator ('and'/'or', not both) — our
    visibility rule only supports one combinator per rule.
    """
    has_and = re.search(r"\band\b", expr, re.IGNORECASE) is not None
    has_or = re.search(r"\bor\b", expr, re.IGNORECASE) is not None
    if has_and and has_or:
        return None, None
    if has_and:
        return [p.strip() for p in re.split(r"\band\b", expr, flags=re.IGNORECASE)], "all"
    if has_or:
        return [p.strip() for p in re.split(r"\bor\b", expr, flags=re.IGNORECASE)], "any"
    return [expr.strip()], "all"


def _parse_condition(part: str) -> dict | None:
    m = _SELECTED_RE.match(part)
    if m:
        return {
            "field_key": m.group("field"),
            "operator": "contains",
            "value": _parse_literal(m.group("value")),
        }
    m = _COND_RE.match(part)
    if m:
        return {
            "field_key": m.group("field"),
            "operator": _COMPARE_OP_MAP[m.group("op")],
            "value": _parse_literal(m.group("value")),
        }
    return None


def convert_relevant(expr: str | None) -> tuple[dict | None, str | None]:
    """`relevant` (skip logic) -> a GeoCore visibility rule. Recognizes
    `${field} = 'value'`, `!=`, `>`, `<`, `>=`, `<=`,
    `selected(${field}, 'value')`, and `and`/`or` chains (not both mixed
    in one expression). Anything else returns (None, warning).
    """
    if not expr or not str(expr).strip():
        return None, None
    expr = str(expr).strip()
    parts, combinator = _split_combinator(expr)
    if parts is None:
        return None, f"relevant expression not converted (mixes 'and'/'or' in one rule): {expr}"
    conditions = []
    for part in parts:
        condition = _parse_condition(part)
        if condition is None:
            return None, f"relevant expression not converted: {expr}"
        conditions.append(condition)
    return {"combinator": combinator, "conditions": conditions}, None


def convert_calculation(expr: str | None) -> tuple[str | None, str | None]:
    """`calculation` -> GeoCore's `{field}`-style expression
    (core/expressions.py). `${field}` references convert straight across;
    functions other than round()/abs()/min()/max()/sum() aren't supported
    by the evaluator, so the calculation still imports but is flagged as a
    warning since it will error until simplified.
    """
    if not expr or not str(expr).strip():
        return None, None
    expr = str(expr).strip()
    converted = _FIELD_REF_RE.sub(lambda m: "{%s}" % m.group(1), expr)
    warning = None
    for m in re.finditer(r"([a-zA-Z_][a-zA-Z0-9_]*)\s*\(", converted):
        if m.group(1) not in _ALLOWED_CALC_FUNCS:
            warning = (
                f"calculation uses unsupported function '{m.group(1)}()' "
                f"and will error until simplified: {expr}"
            )
            break
    return converted, warning


def convert_constraint(expr: str | None) -> tuple[dict | None, str | None]:
    """`constraint` -> a GeoCore validation rule. Only `. >= n`, `. <= n`,
    `. > n`, `. < n` (and `and` chains of those, merged into one min/max
    rule) are recognized — `.` meaning "this question's own answer" is
    the one XLSForm constraint pattern common enough to special-case.
    """
    if not expr or not str(expr).strip():
        return None, None
    expr = str(expr).strip()
    parts, _combinator = _split_combinator(expr)
    if parts is None:
        return None, f"constraint not converted (mixes 'and'/'or' in one rule): {expr}"
    validation: dict[str, float] = {}
    for part in parts:
        m = _CONSTRAINT_RE.match(part.strip())
        if not m:
            return None, f"constraint not converted: {expr}"
        op, num_str = m.groups()
        num = float(num_str) if "." in num_str else int(num_str)
        if op in (">=", ">"):
            validation["min"] = num
        else:
            validation["max"] = num
    return validation, None


# ---------------------------------------------------------------------------
# Row helpers
# ---------------------------------------------------------------------------


def _get(row: dict, *keys: str):
    for key in keys:
        value = row.get(key)
        if value not in (None, ""):
            return value
    lowered = {str(k).lower(): v for k, v in row.items()}
    for key in keys:
        value = lowered.get(key.lower())
        if value not in (None, ""):
            return value
    return None


def _get_label(row: dict) -> str | None:
    value = _get(row, "label")
    if value not in (None, ""):
        return str(value).strip()
    # Multi-language forms use columns like "label::English (en)" — take
    # the first one we find rather than requiring a specific language.
    for key, value in row.items():
        if key and str(key).strip().lower().startswith("label") and value not in (None, ""):
            return str(value).strip()
    return None


def _is_truthy(value: Any) -> bool:
    if value in (None, ""):
        return False
    return str(value).strip().lower() in ("yes", "true", "true()", "1")


def _unique_key(base: str, used: set[str]) -> str:
    key = base
    suffix = 1
    while key in used:
        suffix += 1
        key = f"{base}_{suffix}"
    used.add(key)
    return key


# ---------------------------------------------------------------------------
# Row -> ParsedForm
# ---------------------------------------------------------------------------


def build_form_from_rows(
    survey_rows: list[dict],
    choices_rows: list[dict] | None = None,
    settings_rows: list[dict] | None = None,
) -> ParsedForm:
    choices_rows = choices_rows or []
    settings_rows = settings_rows or []
    warnings: list[str] = []

    form_name = None
    if settings_rows:
        form_name = _get(settings_rows[0], "form_title", "form_id")
    form_name = str(form_name).strip() if form_name else "Imported Form"

    choices_map: dict[str, list[str]] = {}
    for choice_row in choices_rows:
        list_name = _get(choice_row, "list_name", "list name")
        label = _get_label(choice_row)
        if not list_name or not label:
            continue
        choices_map.setdefault(str(list_name).strip(), []).append(label)

    geometry_type: str | None = None
    sections: list[ParsedSection] = []
    stack: list[ParsedSection] = []
    default_section: ParsedSection | None = None
    nested_depth = 0
    used_field_keys: set[str] = set()

    def current_section() -> ParsedSection:
        nonlocal default_section
        if stack:
            return stack[-1]
        if default_section is None:
            default_section = ParsedSection(title="General")
            sections.append(default_section)
        return default_section

    for row in survey_rows:
        raw_type = _get(row, "type")
        if not raw_type:
            continue
        raw_type = str(raw_type).strip()
        type_lower = raw_type.lower()
        norm_type = re.sub(r"\s+", "_", type_lower)

        name = _get(row, "name")
        name = str(name).strip() if name else None
        label = _get_label(row) or name

        if norm_type in ("begin_group", "begin_repeat"):
            if stack:
                nested_depth += 1
                warnings.append(
                    f"Nested group/repeat '{label or name or ''}' was flattened into its "
                    "parent section — GeoCore's form model is one section deep."
                )
                continue
            stack.append(
                ParsedSection(
                    title=label or "Section",
                    repeatable=(norm_type == "begin_repeat"),
                    repeat_label=label if norm_type == "begin_repeat" else None,
                )
            )
            continue

        if norm_type in ("end_group", "end_repeat"):
            if nested_depth > 0:
                nested_depth -= 1
                continue
            if stack:
                sections.append(stack.pop())
            continue

        if norm_type in _DROPPED_TYPES:
            continue

        if norm_type in GEOMETRY_QUESTION_TYPES:
            if geometry_type is None:
                geometry_type = GEOMETRY_QUESTION_TYPES[norm_type]
            else:
                warnings.append(
                    f"Additional location question '{name}' ignored — an asset type has one geometry type."
                )
            continue

        if not name:
            warnings.append(f"Question '{label or raw_type}' has no name column and was skipped.")
            continue

        base_type = raw_type.split()[0].lower()
        options: list[str] | None = None
        if base_type in ("select_one", "select_multiple"):
            list_name = raw_type.split()[1] if len(raw_type.split()) > 1 else None
            if list_name:
                options = choices_map.get(list_name)
                if options is None:
                    warnings.append(
                        f"Choice list '{list_name}' for question '{name}' wasn't found in the choices sheet."
                    )
            else:
                warnings.append(f"Question '{name}' is a select type with no choice list named.")

        field_type = _TYPE_MAP.get(base_type)
        if field_type is None:
            warnings.append(f"Question type '{raw_type}' for '{name}' isn't supported and was skipped.")
            continue

        field_key = name if _IDENTIFIER_RE.match(name) else slugify_key(name)
        field_key = _unique_key(field_key, used_field_keys)

        visibility, vis_warning = convert_relevant(_get(row, "relevant"))
        if vis_warning:
            warnings.append(f"[{name}] {vis_warning}")

        calculation, calc_warning = convert_calculation(_get(row, "calculation"))
        if calc_warning:
            warnings.append(f"[{name}] {calc_warning}")

        validation, val_warning = convert_constraint(_get(row, "constraint"))
        if val_warning:
            warnings.append(f"[{name}] {val_warning}")

        current_section().fields.append(
            ParsedField(
                label=label or name,
                field_type=field_type,
                field_key=field_key,
                options=options,
                is_required=_is_truthy(_get(row, "required")),
                visibility=visibility,
                calculation=calculation,
                validation=validation,
            )
        )

    while stack:
        finished = stack.pop()
        warnings.append(
            f"Section '{finished.title}' was missing its end group/end repeat — closed automatically."
        )
        sections.append(finished)

    if not sections:
        sections = [ParsedSection(title="General")]

    return ParsedForm(
        name=form_name,
        geometry_type=geometry_type or "point",
        sections=sections,
        warnings=warnings,
    )


# ---------------------------------------------------------------------------
# .xlsx -> rows
# ---------------------------------------------------------------------------


def _read_sheet_rows(workbook, sheet_names: list[str]) -> list[dict]:
    sheet = None
    lowered_target = {n.lower() for n in sheet_names}
    for sheet_name in workbook.sheetnames:
        if sheet_name.strip().lower() in lowered_target:
            sheet = workbook[sheet_name]
            break
    if sheet is None:
        return []

    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []
    headers = [str(h).strip() if h is not None else "" for h in header_row]

    rows: list[dict] = []
    for raw_row in rows_iter:
        if raw_row is None or all(v is None for v in raw_row):
            continue
        row = {header: value for header, value in zip(headers, raw_row) if header}
        if row:
            rows.append(row)
    return rows


def parse_xlsform(content: bytes, filename: str | None = None) -> ParsedForm:
    """Reads an uploaded .xlsx's `survey` (required), `choices`, and
    `settings` sheets and converts them into a ParsedForm. Raises
    XLSFormError for anything that stops the import outright (unreadable
    file, missing/empty survey sheet); individual unconvertible rows are
    skipped with a warning instead, never raised.
    """
    try:
        import openpyxl
    except ImportError as exc:
        raise XLSFormError(
            "openpyxl isn't installed on the server — run `pip install -r requirements.txt` and restart."
        ) from exc

    try:
        workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:  # openpyxl raises several different error types for bad files
        raise XLSFormError(
            f"Couldn't read '{filename or 'the uploaded file'}' as an Excel file: {exc}"
        ) from exc

    survey_rows = _read_sheet_rows(workbook, ["survey"])
    if not survey_rows:
        raise XLSFormError(
            "No 'survey' sheet found (or it's empty) — an XLSForm needs a 'survey' sheet "
            "with at least type/name/label columns."
        )

    choices_rows = _read_sheet_rows(workbook, ["choices"])
    settings_rows = _read_sheet_rows(workbook, ["settings"])

    return build_form_from_rows(survey_rows, choices_rows, settings_rows)
